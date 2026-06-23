"""
exporter.py
Manufacturing SPC Dashboard — Export Layer

SPC-specific export *config* (which fields, metadata, layout) composed over the
shared, app-agnostic primitives in ``quality_core.io.export`` (CSV/formula-injection
escaping, openpyxl styling, fpdf2 table rendering). The cross-cutting machinery is
written once in core and reused here, exactly as the FMEA exporter does.

Two report kinds, each rendered to Excel (.xlsx) and PDF:

    Control chart report  — points, control limits, rule violations, chart metrics
    Capability report     — Cp/Cpk/Pp/Ppk, distribution summary, normality, stability

Builders take a frozen report dataclass (the values a page has already computed) and
return raw bytes suitable for ``st.download_button``.
"""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd
from quality_core.io.export import (
    pdf_subheader,
    pdf_summary_cells,
    pdf_title,
    render_table,
    safe_text,
    sanitize_cell,
    sanitize_for_export,
    write_keyvalue_sheet,
    write_table_sheet,
)

from spc_app import __version__

_TOOL_VERSION = __version__
_ENGINEERING_REF = "AIAG SPC Reference Manual, 4th Ed. (2005)"

# Cpk capability bands — mirror the Capability page's interpretation table, which
# cites docs/ASSUMPTIONS_LOG.md (1.33 = common minimum target; 1.00 = not capable).
_CPK_CAPABLE = 1.33
_CPK_MARGINAL = 1.00

_VIOLATION_FILL_HEX = "F8D7DA"  # light red for out-of-control rows
_VIOLATION_RGB = (248, 215, 218)
_WHITE_RGB = (255, 255, 255)


# ===========================================================================
# Report inputs
# ===========================================================================


@dataclass(frozen=True)
class ControlChartReport:
    """Everything the Control Charts page already computed for one chart."""

    chart_label: str
    stream: str
    rule_set: str
    points: Sequence[float]
    cl: float
    # p- and u-charts have per-point (vector) control limits; the others are scalar.
    ucl: float | Sequence[float]
    lcl: float | Sequence[float]
    violations: Sequence[Mapping[str, Any]]  # [{"index": int, "rule": str}, ...]
    metrics: Sequence[tuple[str, str]]  # summarize_metrics() output


@dataclass(frozen=True)
class CapabilityReport:
    """Everything the Process Capability page already computed for one stream."""

    stream_label: str
    values: Sequence[float]
    capability: Mapping[str, Any]  # cp/cpk/pp/ppk/mean/sigma_hat/sigma_overall
    lsl: float | None
    usl: float | None
    normality: Mapping[str, Any]  # is_normal/p_value/w_stat
    oos_signal_count: int


# ===========================================================================
# Formatting helpers
# ===========================================================================


def _fmt(value: float) -> str:
    return f"{value:.4f}"


def _fmt_opt(value: float | None) -> str:
    return "N/A" if value is None else f"{value:.4f}"


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _generated_line(detail: str) -> str:
    """The 'Generated: <timestamp> | <detail>' caption for a PDF sub-header."""
    return f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   {detail}"


def _limit_at(limit: float | Sequence[float], index: int) -> float:
    """Resolve a control limit at a point: a scalar applies to every point, a
    vector (p/u charts) gives the per-point limit."""
    if isinstance(limit, (int, float)):
        return float(limit)
    return float(limit[index])


def _fmt_limit(limit: float | Sequence[float]) -> str:
    """A scalar limit prints as a number; a vector limit varies per subgroup."""
    if isinstance(limit, (int, float)):
        return _fmt(float(limit))
    return "varies (per subgroup)"


def _cpk_rating(cpk: float | None) -> str:
    if cpk is None:
        return "N/A"
    if cpk >= _CPK_CAPABLE:
        return "Capable"
    if cpk >= _CPK_MARGINAL:
        return "Marginal"
    return "Not capable"


def _violations_by_index(report: ControlChartReport) -> dict[int, list[str]]:
    by_index: dict[int, list[str]] = {}
    for violation in report.violations:
        by_index.setdefault(int(violation["index"]), []).append(str(violation["rule"]))
    return by_index


_POINT_COLUMNS = ["Point", "Value", "UCL", "LCL", "Status"]


def _points_frame(report: ControlChartReport) -> pd.DataFrame:
    """Per-point table: 1-based Point, Value, the UCL/LCL it was tested against
    (constant for most charts, per-point for p/u), and OK / rule-violation Status."""
    by_index = _violations_by_index(report)
    rows = [
        {
            "Point": index + 1,
            "Value": round(float(value), 6),
            "UCL": round(_limit_at(report.ucl, index), 6),
            "LCL": round(_limit_at(report.lcl, index), 6),
            "Status": "; ".join(by_index[index]) if index in by_index else "OK",
        }
        for index, value in enumerate(report.points)
    ]
    return pd.DataFrame(rows, columns=_POINT_COLUMNS)


def _values_frame(values: Sequence[float]) -> pd.DataFrame:
    rows = [{"Point": i + 1, "Value": round(float(v), 6)} for i, v in enumerate(values)]
    return pd.DataFrame(rows, columns=["Point", "Value"])


# ===========================================================================
# Control chart report
# ===========================================================================


def _control_chart_summary_rows(report: ControlChartReport) -> list[tuple[str, object]]:
    rows: list[tuple[str, object]] = [
        ("Generated", _now()),
        ("Tool Version", _TOOL_VERSION),
        ("Engineering Ref", _ENGINEERING_REF),
        ("", ""),
        ("Chart Type", sanitize_cell(report.chart_label)),
        ("Process Stream", sanitize_cell(report.stream)),
        ("Rule Set", sanitize_cell(report.rule_set)),
        ("Center Line (CL)", _fmt(report.cl)),
        ("UCL", _fmt_limit(report.ucl)),
        ("LCL", _fmt_limit(report.lcl)),
        ("Data Points", len(report.points)),
        ("Rule Violations", len(report.violations)),
        ("", ""),
    ]
    # Metrics are app-formatted numeric strings (e.g. "10.0000", "-3.0000"); they are
    # NOT user input, so they bypass sanitize_cell — escaping them would prefix a
    # spurious apostrophe onto legitimate negative values.
    rows += list(report.metrics)
    if report.violations:
        rows.append(("", ""))
        for violation in report.violations:
            point = int(violation["index"]) + 1
            rows.append((f"Violation @ Point {point}", sanitize_cell(str(violation["rule"]))))
    return rows


def _control_chart_row_fill(row: pd.Series) -> str | None:
    return _VIOLATION_FILL_HEX if str(row.get("Status")) != "OK" else None


def build_control_chart_report_excel(report: ControlChartReport) -> bytes:
    """Excel workbook: a coloured per-point sheet + a summary/metadata sheet."""
    points = sanitize_for_export(_points_frame(report))
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None  # a freshly created workbook always has an active sheet
    write_table_sheet(
        ws,
        points,
        title="Control Chart",
        columns=_POINT_COLUMNS,
        col_widths={"Point": 8, "Value": 14, "UCL": 14, "LCL": 14, "Status": 40},
        row_fill_hex=_control_chart_row_fill,
    )
    write_keyvalue_sheet(wb.create_sheet("Summary"), _control_chart_summary_rows(report))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_control_chart_report_pdf(report: ControlChartReport) -> bytes:
    """PDF report: title, summary metric strip, and a per-point violations table."""
    from fpdf import FPDF

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    pdf_title(pdf, "SPC Control Chart Report")
    pdf_subheader(pdf, _generated_line(f"{report.chart_label}  |  {_ENGINEERING_REF}"))
    pdf_summary_cells(
        pdf,
        [
            ("CL", _fmt(report.cl)),
            ("UCL", _fmt_limit(report.ucl)),
            ("LCL", _fmt_limit(report.lcl)),
            ("Points", str(len(report.points))),
            ("Violations", str(len(report.violations))),
        ],
    )

    render_table(
        pdf,
        _points_frame(report),
        columns=[("Point", 18), ("Value", 30), ("UCL", 30), ("LCL", 30), ("Status", 82)],
        row_values=lambda r: [
            safe_text(str(r["Point"])),
            safe_text(str(r["Value"])),
            safe_text(str(r["UCL"])),
            safe_text(str(r["LCL"])),
            safe_text(str(r["Status"])),
        ],
        row_rgb=lambda r: _VIOLATION_RGB if str(r["Status"]) != "OK" else _WHITE_RGB,
    )
    return bytes(pdf.output())


# ===========================================================================
# Capability report
# ===========================================================================


def _capability_detail_rows(report: CapabilityReport) -> list[tuple[str, object]]:
    cap = report.capability
    norm = report.normality
    stability = (
        "In statistical control"
        if report.oos_signal_count == 0
        else f"{report.oos_signal_count} WE signal(s) — Cpk indicative only"
    )
    return [
        ("Process Stream", sanitize_cell(report.stream_label)),
        ("Data Points", len(report.values)),
        ("LSL", _fmt_opt(report.lsl)),
        ("USL", _fmt_opt(report.usl)),
        ("Cp", _fmt_opt(cap["cp"])),
        ("Cpk", _fmt_opt(cap["cpk"])),
        ("Pp", _fmt_opt(cap["pp"])),
        ("Ppk", _fmt_opt(cap["ppk"])),
        ("Cpk Rating", _cpk_rating(cap["cpk"])),
        ("Mean", _fmt(cap["mean"])),
        ("Sigma Hat (within)", _fmt(cap["sigma_hat"])),
        ("Sigma Overall", _fmt(cap["sigma_overall"])),
        ("Normality (Shapiro-Wilk p)", _fmt(norm["p_value"])),
        ("Approximately Normal?", "Yes" if norm["is_normal"] else "No"),
        ("Stability", stability),
    ]


def _capability_summary_rows(report: CapabilityReport) -> list[tuple[str, object]]:
    return [
        ("Generated", _now()),
        ("Tool Version", _TOOL_VERSION),
        ("Engineering Ref", _ENGINEERING_REF),
        ("", ""),
        *_capability_detail_rows(report),
    ]


def build_capability_report_excel(report: CapabilityReport) -> bytes:
    """Excel workbook: a capability/metadata summary sheet + the raw data sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    write_keyvalue_sheet(ws, _capability_summary_rows(report), title="Capability")
    write_table_sheet(
        wb.create_sheet("Data"),
        sanitize_for_export(_values_frame(report.values)),
        title="Data",
        columns=["Point", "Value"],
        col_widths={"Point": 8, "Value": 16},
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_capability_report_pdf(report: CapabilityReport) -> bytes:
    """PDF report: title, Cp/Cpk/Pp/Ppk strip, and a capability detail table."""
    from fpdf import FPDF

    cap = report.capability
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    pdf_title(pdf, "SPC Process Capability Report")
    pdf_subheader(pdf, _generated_line(f"{report.stream_label}  |  {_ENGINEERING_REF}"))
    pdf_summary_cells(
        pdf,
        [
            ("Cp", _fmt_opt(cap["cp"])),
            ("Cpk", _fmt_opt(cap["cpk"])),
            ("Pp", _fmt_opt(cap["pp"])),
            ("Ppk", _fmt_opt(cap["ppk"])),
        ],
    )

    details = pd.DataFrame(_capability_detail_rows(report), columns=["Metric", "Value"])
    render_table(
        pdf,
        details,
        columns=[("Metric", 90), ("Value", 100)],
        row_values=lambda r: [safe_text(str(r["Metric"])), safe_text(str(r["Value"]))],
        row_rgb=lambda r: _WHITE_RGB,
    )
    return bytes(pdf.output())
