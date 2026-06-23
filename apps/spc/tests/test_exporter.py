"""Tests for spc_app/exporter.py — SPC report export (W04-3).

SPC-specific export config composed over quality_core.io.export. These tests load
the produced .xlsx back with openpyxl and assert structure, colouring, and that
user-derived strings are formula-injection-escaped; PDFs are checked for valid
bytes. The shared export machinery itself is tested in quality-core.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from spc_app.exporter import (
    CapabilityReport,
    ControlChartReport,
    _cpk_rating,
    _fmt_opt,
    _points_frame,
    build_capability_report_excel,
    build_capability_report_pdf,
    build_control_chart_report_excel,
    build_control_chart_report_pdf,
)

# --- Fixtures ----------------------------------------------------------------


def _control_chart_report(stream: str = "ply_thickness") -> ControlChartReport:
    return ControlChartReport(
        chart_label="Xbar-R Chart",
        stream=stream,
        rule_set="Western Electric",
        points=[10.0, 10.5, 13.9, 9.8, 10.1],
        cl=10.0,
        ucl=12.0,
        lcl=8.0,
        violations=[{"index": 2, "rule": "Rule 1: beyond 3-sigma"}],
        metrics=[("Xbarbar", "10.0000"), ("Rbar", "1.2000"), ("Sigma Hat", "0.5000")],
    )


def _capability_report(stream_label: str = "Ply Thickness") -> CapabilityReport:
    return CapabilityReport(
        stream_label=stream_label,
        values=[10.0, 10.1, 9.9, 10.2, 9.8, 10.05],
        capability={
            "cp": 1.45,
            "cpk": 1.21,
            "pp": 1.40,
            "ppk": 1.18,
            "mean": 10.0,
            "sigma_hat": 0.05,
            "sigma_overall": 0.06,
        },
        lsl=9.7,
        usl=10.3,
        normality={"w_stat": 0.98, "p_value": 0.61, "is_normal": True},
        oos_signal_count=0,
    )


def _kv_sheet_to_dict(ws) -> dict[str, object]:
    return {row[0].value: row[1].value for row in ws.iter_rows(min_row=1, max_col=2) if row[0].value}


# --- pure helpers ------------------------------------------------------------


def test_points_frame_marks_violations_and_is_one_based():
    frame = _points_frame(_control_chart_report())
    assert list(frame["Point"]) == [1, 2, 3, 4, 5]
    statuses = list(frame["Status"])
    assert statuses[2] == "Rule 1: beyond 3-sigma"  # index 2 → point 3
    assert all(s == "OK" for i, s in enumerate(statuses) if i != 2)
    # Scalar limits apply to every point.
    assert list(frame["UCL"]) == [12.0] * 5
    assert list(frame["LCL"]) == [8.0] * 5


def test_points_frame_handles_per_point_vector_limits():
    # p/u charts carry per-point UCL/LCL vectors; each point shows its own limit.
    report = ControlChartReport(
        chart_label="p Chart",
        stream="reject_proportion",
        rule_set="Western Electric",
        points=[0.10, 0.20, 0.15],
        cl=0.15,
        ucl=[0.25, 0.30, 0.28],
        lcl=[0.05, 0.00, 0.02],
        violations=[],
        metrics=[("pbar", "0.1500")],
    )
    frame = _points_frame(report)
    assert list(frame["UCL"]) == [0.25, 0.30, 0.28]
    assert list(frame["LCL"]) == [0.05, 0.00, 0.02]


def test_cpk_rating_bands():
    assert _cpk_rating(None) == "N/A"
    assert _cpk_rating(0.9) == "Not capable"
    assert _cpk_rating(1.0) == "Marginal"
    assert _cpk_rating(1.32) == "Marginal"
    assert _cpk_rating(1.33) == "Capable"
    assert _cpk_rating(2.0) == "Capable"


def test_fmt_opt_handles_none():
    assert _fmt_opt(None) == "N/A"
    assert _fmt_opt(1.5) == "1.5000"


# --- control chart Excel -----------------------------------------------------


def test_control_chart_excel_structure_and_violation_colour():
    data = build_control_chart_report_excel(_control_chart_report())
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Control Chart", "Summary"]

    ws = wb["Control Chart"]
    assert [c.value for c in ws[1]] == ["Point", "Value", "UCL", "LCL", "Status"]
    # The violating point (row 4 = point 3) carries the red fill; an OK row does not.
    assert ws.cell(row=4, column=5).value == "Rule 1: beyond 3-sigma"
    assert ws.cell(row=4, column=1).fill.fgColor.rgb.endswith("F8D7DA")
    assert ws.cell(row=2, column=1).fill.fill_type in (None, "none")

    summary = _kv_sheet_to_dict(wb["Summary"])
    assert summary["Chart Type"] == "Xbar-R Chart"
    assert summary["UCL"] == "12.0000"
    assert summary["Rule Violations"] == 1
    assert summary["Tool Version"]  # version stamped


def test_control_chart_excel_summary_reports_varying_limits():
    report = ControlChartReport(
        chart_label="p Chart",
        stream="reject_proportion",
        rule_set="Nelson",
        points=[0.1, 0.2],
        cl=0.15,
        ucl=[0.25, 0.30],
        lcl=[0.05, 0.00],
        violations=[],
        metrics=[("pbar", "0.1500")],
    )
    wb = openpyxl.load_workbook(io.BytesIO(build_control_chart_report_excel(report)))
    summary = _kv_sheet_to_dict(wb["Summary"])
    assert summary["UCL"] == "varies (per subgroup)"


def test_control_chart_excel_does_not_escape_negative_metric_values():
    # Metrics are app-formatted numbers, not user input; a negative value like
    # "-3.0000" must NOT gain a spurious leading apostrophe in the summary sheet.
    report = ControlChartReport(
        chart_label="I-MR Chart",
        stream="autoclave_temp",
        rule_set="Western Electric",
        points=[-3.0, -2.5],
        cl=-2.75,
        ucl=0.0,
        lcl=-5.5,
        violations=[],
        metrics=[("Xbar", "-2.7500"), ("MRbar", "0.5000")],
    )
    wb = openpyxl.load_workbook(io.BytesIO(build_control_chart_report_excel(report)))
    summary = _kv_sheet_to_dict(wb["Summary"])
    assert summary["Xbar"] == "-2.7500"  # not "'-2.7500"


def test_control_chart_excel_escapes_injection_in_stream_name():
    # An uploaded stream name beginning with a formula trigger must be escaped.
    data = build_control_chart_report_excel(_control_chart_report(stream="=cmd|' /c calc'!A1"))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    summary = _kv_sheet_to_dict(wb["Summary"])
    assert str(summary["Process Stream"]).startswith("'=")


def test_control_chart_pdf_is_valid_bytes():
    data = build_control_chart_report_pdf(_control_chart_report())
    assert data.startswith(b"%PDF")
    assert len(data) > 1000


# --- capability Excel --------------------------------------------------------


def test_capability_excel_structure_and_values():
    data = build_capability_report_excel(_capability_report())
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Capability", "Data"]

    summary = _kv_sheet_to_dict(wb["Capability"])
    assert summary["Cpk"] == "1.2100"
    assert summary["Cpk Rating"] == "Marginal"
    assert summary["Stability"] == "In statistical control"

    data_ws = wb["Data"]
    assert [c.value for c in data_ws[1]] == ["Point", "Value"]
    assert data_ws.max_row == 1 + 6  # header + 6 values


def test_capability_excel_escapes_injection_in_stream_label():
    data = build_capability_report_excel(_capability_report(stream_label="=HYPERLINK(0)"))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    summary = _kv_sheet_to_dict(wb["Capability"])
    assert str(summary["Process Stream"]).startswith("'=")


def test_capability_excel_reports_na_when_limits_missing():
    report = _capability_report()
    report = CapabilityReport(
        stream_label=report.stream_label,
        values=report.values,
        capability={**report.capability, "cp": None, "cpk": None, "pp": None, "ppk": None},
        lsl=None,
        usl=None,
        normality=report.normality,
        oos_signal_count=0,
    )
    wb = openpyxl.load_workbook(io.BytesIO(build_capability_report_excel(report)))
    summary = _kv_sheet_to_dict(wb["Capability"])
    assert summary["LSL"] == "N/A"
    assert summary["Cpk"] == "N/A"
    assert summary["Cpk Rating"] == "N/A"


def test_capability_excel_flags_out_of_control():
    report = _capability_report()
    report = CapabilityReport(
        stream_label=report.stream_label,
        values=report.values,
        capability=report.capability,
        lsl=report.lsl,
        usl=report.usl,
        normality=report.normality,
        oos_signal_count=3,
    )
    wb = openpyxl.load_workbook(io.BytesIO(build_capability_report_excel(report)))
    summary = _kv_sheet_to_dict(wb["Capability"])
    assert "3 WE signal(s)" in str(summary["Stability"])


def test_capability_pdf_is_valid_bytes():
    data = build_capability_report_pdf(_capability_report())
    assert data.startswith(b"%PDF")
    assert len(data) > 1000


@pytest.mark.parametrize("oos", [0, 2])
def test_capability_pdf_handles_stability_states(oos):
    report = _capability_report()
    report = CapabilityReport(
        stream_label=report.stream_label,
        values=report.values,
        capability=report.capability,
        lsl=report.lsl,
        usl=report.usl,
        normality={"w_stat": 0.9, "p_value": 0.01, "is_normal": False},
        oos_signal_count=oos,
    )
    assert build_capability_report_pdf(report).startswith(b"%PDF")
