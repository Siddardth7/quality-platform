"""
quality_core/io/export.py
Shared export primitives for Quality Platform apps — CSV, Excel (openpyxl), PDF (fpdf2).

App-agnostic by design: no FMEA/SPC domain knowledge lives here. Apps supply the
columns, widths, row colors, and PDF layout; these helpers own the cross-cutting
concerns that every app's export must get right:

  - formula-injection escaping for CSV/Excel (`sanitize_for_export`, `export_csv`)
  - openpyxl table + key/value sheet styling (`write_table_sheet`, `write_keyvalue_sheet`)
  - Latin-1 text sanitizing for fpdf2 core fonts (`safe_text`)
  - a repeating-header, page-breaking PDF table body (`render_table`) and image page
    embedding (`add_image_page`)

This is the platform's "write export once, use it everywhere" surface: FMEA composes
these today; SPC and the Control Plan reuse them verbatim.
"""

from __future__ import annotations

from collections.abc import Callable, Mapping, Sequence
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ===========================================================================
# Formula-injection escaping (CSV / Excel)
# ===========================================================================

#: OWASP CSV-injection trigger characters. A cell whose leading character is one
#: of these can be evaluated by a spreadsheet as a formula/command — including the
#: Tab (0x09) and CR (0x0D) controls, which several spreadsheet apps treat as
#: formula leads. The escape character (apostrophe) must NOT be added here — that
#: would double-escape on a repeated call.
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _is_injection_risk(value: str) -> bool:
    """True if ``value`` could be evaluated as a formula by a spreadsheet."""
    if value.startswith(FORMULA_PREFIXES):
        return True
    # Excel/Sheets strip leading whitespace before formula detection, so a formula
    # character hidden behind leading spaces/tabs/newlines is still dangerous.
    stripped = value.lstrip()
    return stripped != value and stripped.startswith(("=", "+", "-", "@"))


def sanitize_for_export(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy with formula-injection-risky string cells escaped.

    A string whose first non-whitespace character is a formula trigger (``= + - @``,
    or a leading Tab/CR) is prefixed with ``'`` so Excel/Sheets/LibreOffice render it
    literally instead of evaluating it. Non-string cells are untouched, and the escape
    is idempotent (an already-escaped value is not re-escaped).
    """
    df = df.copy()
    for col in df.columns:
        df[col] = df[col].apply(
            lambda v: f"'{v}" if isinstance(v, str) and _is_injection_risk(v) else v
        )
    return df


def export_csv(df: pd.DataFrame) -> bytes:
    """Export a DataFrame to UTF-8 CSV bytes with formula-injection escaping."""
    return sanitize_for_export(df).to_csv(index=False).encode("utf-8")


# ===========================================================================
# Excel styling (openpyxl)
# ===========================================================================

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)


def write_table_sheet(
    ws: Any,
    df: pd.DataFrame,
    *,
    title: str,
    columns: Sequence[str],
    col_widths: Mapping[str, float],
    row_fill_hex: Callable[[pd.Series], str | None] | None = None,
    header_fill_hex: str = "2C3E50",
    default_width: float = 14,
    header_height: float = 22,
    freeze: str | None = "A2",
) -> None:
    """Write ``df`` as a styled table into worksheet ``ws``.

    Only ``columns`` present in ``df`` are written, in the given order. ``row_fill_hex``
    maps a row to a solid fill color (hex, no ``#``) or ``None`` for no fill — apps use
    it to colour rows by risk tier. numpy scalars are unwrapped via ``.item()`` for
    openpyxl compatibility. Fills are cached per color so repeated colours produce
    identical ``PatternFill`` objects.
    """
    ws.title = title
    cols = [c for c in columns if c in df.columns]

    header_fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = header_height

    fill_cache: dict[str, PatternFill] = {}
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        hex_val = row_fill_hex(row) if row_fill_hex is not None else None
        fill = None
        if hex_val is not None:
            fill = fill_cache.get(hex_val)
            if fill is None:
                fill = PatternFill(start_color=hex_val, end_color=hex_val, fill_type="solid")
                fill_cache[hex_val] = fill
        for col_idx, col_name in enumerate(cols, start=1):
            val = row[col_name]
            if hasattr(val, "item"):  # numpy bool/int -> native
                val = val.item()
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill is not None:
                cell.fill = fill
            cell.font = _NORMAL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    for col_idx, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, default_width)

    if freeze:
        ws.freeze_panes = freeze


def write_keyvalue_sheet(
    ws: Any,
    rows: Sequence[tuple[str, object]],
    *,
    key_width: float = 22,
    value_width: float = 48,
) -> None:
    """Write a two-column (bold key, plain value) metadata sheet into ``ws``."""
    for r_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(r_idx, 1, label).font = _BOLD_FONT
        ws.cell(r_idx, 2, value).font = _NORMAL_FONT
    ws.column_dimensions["A"].width = key_width
    ws.column_dimensions["B"].width = value_width


# ===========================================================================
# PDF text + table (fpdf2)
# ===========================================================================

# fpdf2 core fonts only support Latin-1; map the common Unicode we emit to ASCII.
_UNICODE_MAP = [
    ("—", "-"),    # em dash  —
    ("–", "-"),    # en dash  –
    ("×", "x"),    # multiplication sign  ×
    ("≥", ">="),   # ≥
    ("≤", "<="),   # ≤
    ("±", "+/-"),  # ±
    ("°", " deg"), # °
    ("‘", "'"),    # left single quote
    ("’", "'"),    # right single quote
    ("“", '"'),    # left double quote
    ("”", '"'),    # right double quote
    ("•", "*"),    # bullet
    ("é", "e"),    # é
    ("à", "a"),    # à
]


def safe_text(s: object) -> str:
    """Sanitize a value to Latin-1 safe text for fpdf2 core fonts."""
    text = str(s)
    for char, rep in _UNICODE_MAP:
        text = text.replace(char, rep)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def render_table(
    pdf: Any,
    df: pd.DataFrame,
    *,
    columns: Sequence[tuple[str, float]],
    row_values: Callable[[pd.Series], Sequence[str]],
    row_rgb: Callable[[pd.Series], tuple[int, int, int]],
    header_fill_rgb: tuple[int, int, int] = (44, 62, 80),
    header_font_size: float = 8,
    row_font_size: float = 7,
    row_height: float = 6,
) -> None:
    """Render ``df`` as a bordered table into the current ``pdf`` position.

    ``columns`` is a sequence of (label, width_mm). The header row is repeated
    whenever a page break is needed (less than two rows of space remain). ``row_values``
    yields the per-row cell strings (same length/order as ``columns``); ``row_rgb`` gives
    each row's fill colour. Cells ≤16 mm wide are centre-aligned, wider ones left-aligned.
    """

    def _header() -> None:
        pdf.set_font("Helvetica", "B", header_font_size)
        pdf.set_fill_color(*header_fill_rgb)
        pdf.set_text_color(255, 255, 255)
        for label, width in columns:
            pdf.cell(width, 7, label, border=1, align="C", fill=True)
        pdf.ln()

    _header()
    pdf.set_font("Helvetica", "", row_font_size)

    for _, row in df.iterrows():
        # Page-break guard: repeat the header when less than 2 rows of space remain.
        if pdf.get_y() + row_height * 2 > pdf.h - pdf.b_margin:
            pdf.add_page()
            _header()
            pdf.set_font("Helvetica", "", row_font_size)

        r, g, b = row_rgb(row)
        pdf.set_fill_color(r, g, b)
        pdf.set_text_color(40, 40, 40)
        for (_, width), val in zip(columns, row_values(row), strict=True):
            pdf.cell(width, row_height, val, border=1, align="C" if width <= 16 else "L", fill=True)
        pdf.ln()


def add_image_page(
    pdf: Any,
    png_path: str,
    title: str,
    *,
    header_fill_rgb: tuple[int, int, int] = (44, 62, 80),
    image_width: float = 277,
) -> None:
    """Add a new landscape page with a title bar and a full-width embedded PNG."""
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_fill_color(*header_fill_rgb)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 9, safe_text(title), new_x="LMARGIN", new_y="NEXT", align="C", fill=True)
    pdf.ln(4)
    pdf.image(png_path, x=10, w=image_width)
