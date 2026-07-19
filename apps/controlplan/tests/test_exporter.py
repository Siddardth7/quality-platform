"""Tests for controlplan_app/exporter.py — CSV/Excel/PDF export of a validated
ControlPlanDataset, mirroring apps/fmea/tests/test_exporter.py's coverage shape
(happy path, empty-dataset guard, formula-injection escaping).
"""

from __future__ import annotations

import io

import openpyxl
from controlplan_app.exporter import export_csv, export_excel, export_pdf
from controlplan_app.schema import ControlPlanDataset, ControlPlanRow


def _row(**overrides) -> ControlPlanRow:
    defaults = dict(
        characteristic="Bore Diameter",
        lsl=9.9,
        usl=10.1,
        target=10.0,
        measurement_method="Bore gauge",
        sample_size=5,
        frequency="per shift",
        recommended_chart="Xbar-R",
        reaction_plan="Stop line; notify quality engineer.",
    )
    defaults.update(overrides)
    return ControlPlanRow(**defaults)


def _dataset(rows: list[ControlPlanRow] | None = None) -> ControlPlanDataset:
    if rows is None:
        rows = [_row()]
    return ControlPlanDataset(rows=rows)


EMPTY = ControlPlanDataset(rows=[])


# --- CSV ---------------------------------------------------------------------


def test_export_csv_returns_nonempty_bytes():
    out = export_csv(_dataset())
    assert isinstance(out, bytes)
    assert len(out) > 0


def test_export_csv_header_matches_export_order():
    out = export_csv(_dataset())
    header = out.decode("utf-8").splitlines()[0].split(",")
    assert header == [
        "characteristic", "lsl", "usl", "target", "measurement_method",
        "sample_size", "frequency", "recommended_chart", "reaction_plan",
        "source_cause_id",
    ]


def test_export_csv_empty_dataset_does_not_crash():
    out = export_csv(EMPTY)
    assert isinstance(out, bytes)
    header = out.decode("utf-8").splitlines()[0].split(",")
    assert header == [
        "characteristic", "lsl", "usl", "target", "measurement_method",
        "sample_size", "frequency", "recommended_chart", "reaction_plan",
        "source_cause_id",
    ]


def test_export_csv_escapes_formula_injection():
    row = _row(reaction_plan="=cmd()|' /C calc'!A1", characteristic="Panel Skin")
    out = export_csv(_dataset([row])).decode("utf-8")
    # Raw unescaped formula must not appear at a cell boundary.
    assert "\n=cmd" not in out and ",=cmd" not in out
    assert "'=cmd" in out


# --- Excel ---------------------------------------------------------------------


def test_export_excel_returns_nonempty_bytes():
    out = export_excel(_dataset())
    assert isinstance(out, bytes)
    assert len(out) > 0


def test_export_excel_valid_workbook_two_sheets():
    out = export_excel(_dataset())
    wb = openpyxl.load_workbook(io.BytesIO(out))
    assert wb.sheetnames == ["Control Plan", "Metadata"]


def test_export_excel_row_count():
    out = export_excel(_dataset([_row(), _row(characteristic="Panel Skin")]))
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb["Control Plan"]
    assert ws.max_row == 2 + 1  # header + 2 rows


def test_export_excel_metadata_sheet_has_tool_version_and_row_count():
    out = export_excel(_dataset())
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb["Metadata"]
    labels = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert "Tool Version" in labels
    assert "Row Count" in labels


def test_export_excel_empty_dataset_does_not_crash():
    out = export_excel(EMPTY)
    wb = openpyxl.load_workbook(io.BytesIO(out))
    assert wb["Control Plan"].max_row == 1  # header only


def test_export_excel_escapes_formula_injection():
    row = _row(reaction_plan="=cmd()", characteristic="+bad")
    out = export_excel(_dataset([row]))
    wb = openpyxl.load_workbook(io.BytesIO(out), data_only=False)
    ws = wb["Control Plan"]
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            assert cell.data_type != "f", (
                f"Cell {cell.coordinate} stored as a formula: {cell.value}"
            )


# --- PDF ---------------------------------------------------------------------


def test_export_pdf_returns_nonempty_bytes():
    out = export_pdf(_dataset())
    assert isinstance(out, bytes)
    assert len(out) > 0


def test_export_pdf_magic_bytes():
    assert export_pdf(_dataset())[:4] == b"%PDF"


def test_export_pdf_empty_dataset_does_not_crash():
    out = export_pdf(EMPTY)
    assert isinstance(out, bytes)
    assert out[:4] == b"%PDF"


def test_export_pdf_multi_row_does_not_raise():
    rows = [_row(characteristic=f"Char {i}") for i in range(5)]
    out = export_pdf(_dataset(rows))
    assert isinstance(out, bytes)
    assert len(out) > 0


def test_export_pdf_handles_null_optional_fields():
    """lsl/usl/target/recommended_chart are nullable — the row-values builder
    must render them without raising (pd.isna branch)."""
    row = _row(lsl=None, usl=None, target=None, recommended_chart=None)
    out = export_pdf(_dataset([row]))
    assert isinstance(out, bytes) and len(out) > 0


def test_to_dataframe_directly_covers_empty_and_nonempty_branches():
    """Exercise the private helper's both branches explicitly, in case a
    future dataset-only exporter usage skips the public functions above."""
    from controlplan_app.exporter import _to_dataframe

    empty_df = _to_dataframe(EMPTY)
    assert list(empty_df.columns) == [
        "characteristic", "lsl", "usl", "target", "measurement_method",
        "sample_size", "frequency", "recommended_chart", "reaction_plan",
        "source_cause_id",
    ]
    assert empty_df.empty

    nonempty_df = _to_dataframe(_dataset())
    assert len(nonempty_df) == 1
    assert list(nonempty_df.columns) == list(empty_df.columns)
