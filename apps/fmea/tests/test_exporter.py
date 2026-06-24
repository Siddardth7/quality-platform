"""
test_exporter.py
Tests for src/exporter.py — Excel and PDF export functions.
"""

import io

import openpyxl
import pandas as pd
import pytest

from fmea_app.ap_engine import calculate_ap
from fmea_app.exporter import export_csv, export_excel, export_pdf


def _sample_df() -> pd.DataFrame:
    """Minimal analyzed FMEA DataFrame (output shape of run_pipeline)."""
    return pd.DataFrame({
        "ID":                     [1, 2, 3],
        "Process_Step":           ["Layup", "Cure", "Demold"],
        "Component":              ["Ply", "Bag", "Part"],
        "Function":               ["F1", "F2", "F3"],
        "Failure_Mode":           ["FM-1", "FM-2", "FM-3"],
        "Effect":                 ["E1", "E2", "E3"],
        "Severity":               [9, 6, 3],
        "Cause":                  ["C1", "C2", "C3"],
        "Occurrence":             [3, 4, 2],
        "Current_Control":        ["Ctrl1", "Ctrl2", "Ctrl3"],
        "Detection":              [4, 5, 2],
        "RPN":                    [108, 120, 12],
        "Risk_Tier":              ["Red", "Red", "Green"],
        "Flag_High_RPN":          [True, True, False],
        "Flag_High_Severity":     [True, False, False],
        "Flag_Action_Priority_H": [True, False, False],
    })


class TestExportExcel:

    def test_returns_bytes(self):
        result = export_excel(_sample_df())
        assert isinstance(result, bytes)

    def test_bytes_non_empty(self):
        result = export_excel(_sample_df())
        assert len(result) > 0

    def test_valid_xlsx_format(self):
        result = export_excel(_sample_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_has_two_sheets(self):
        result = export_excel(_sample_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 2

    def test_sheet_names(self):
        result = export_excel(_sample_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb.sheetnames[0] == "FMEA Analysis"
        assert wb.sheetnames[1] == "Metadata"

    def test_fmea_sheet_row_count(self):
        """Header + 3 data rows = 4 rows total."""
        df = _sample_df()
        result = export_excel(df)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["FMEA Analysis"]
        assert ws.max_row == len(df) + 1  # +1 for header

    def test_fmea_sheet_has_rpn_column(self):
        result = export_excel(_sample_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["FMEA Analysis"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "RPN" in headers

    def test_metadata_sheet_has_content(self):
        result = export_excel(_sample_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Metadata"]
        assert ws.max_row >= 3


# ---------------------------------------------------------------------------
# PDF tests
# ---------------------------------------------------------------------------


class TestExportPdf:

    def test_returns_bytes(self):
        result = export_pdf(_sample_df())
        assert isinstance(result, bytes)

    def test_bytes_non_empty(self):
        result = export_pdf(_sample_df())
        assert len(result) > 5000  # a real PDF is never this small

    def test_pdf_magic_bytes(self):
        """PDF files always start with %PDF."""
        result = export_pdf(_sample_df())
        assert result[:4] == b"%PDF"

    def test_single_row_does_not_raise(self):
        single = _sample_df().head(1)
        result = export_pdf(single)
        assert isinstance(result, bytes)


# ---------------------------------------------------------------------------
# Formula injection tests
# ---------------------------------------------------------------------------

def _pipeline_df_with_formula():
    """DataFrame with formula-injection strings in text fields."""
    import pandas as pd

    from fmea_app.rpn_engine import run_pipeline
    df = pd.DataFrame([{
        "ID": 1, "Process_Step": "=SUM(1,2)",
        "Component": "+malicious", "Function": "Structural support",
        "Failure_Mode": "=2+2", "Effect": "@badcell",
        "Severity": 8, "Cause": "-exploit",
        "Occurrence": 3, "Current_Control": "Visual inspection",
        "Detection": 4,
    }])
    return run_pipeline(df)

def test_excel_no_formula_injection():
    df = _pipeline_df_with_formula()
    raw = export_excel(df)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=False)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            assert cell.data_type != "f", (
                f"Cell {cell.coordinate} was stored as a formula: {cell.value}"
            )

def test_sanitize_escapes_formula_prefixes():
    from fmea_app.exporter import sanitize_for_export
    df = pd.DataFrame([{"Failure_Mode": "=evil", "Component": "+bad", "Cause": "-exploit", "Effect": "'-also", "Process_Step": "@nope", "ID": 1}])
    result = sanitize_for_export(df)
    assert result.loc[0, "Failure_Mode"] == "'=evil"
    assert result.loc[0, "Component"] == "'+bad"
    assert result.loc[0, "Cause"] == "'-exploit"
    assert result.loc[0, "Process_Step"] == "'@nope"


def test_csv_no_formula_injection():
    """CSV export must escape formula-injection prefixes."""
    from fmea_app.exporter import sanitize_for_export
    df = pd.DataFrame([{
        "ID": 1, "Process_Step": "=SUM(1,2)", "Component": "+bad",
        "Function": "Support", "Failure_Mode": "-exploit",
        "Effect": "@nope", "Severity": 8,
        "Cause": "-50C ambient", "Occurrence": 3,
        "Current_Control": "Inspection", "Detection": 4,
    }])
    sanitized = sanitize_for_export(df)
    csv_text = sanitized.to_csv(index=False)
    # Raw (unescaped) formula-starting values must not appear as cell values.
    # A cell value starting with a formula prefix would appear after a comma or newline.
    import re
    for raw_value in ("=SUM(1,2)", "+bad", "-exploit", "@nope"):
        # Match the raw value at a cell boundary (after comma/newline, not preceded by apostrophe)
        pattern = r"(?<!['])(?:(?<=,)|(?<=\n))" + re.escape(raw_value)
        assert not re.search(pattern, csv_text), (
            f"Unescaped formula value '{raw_value}' found in CSV"
        )
    # Escaped versions should be present
    assert "'=SUM(1,2)" in csv_text
    assert "'+bad" in csv_text


def test_excel_includes_ap_column_when_present():
    """W03-3: the AP column is carried into the Excel FMEA sheet header."""
    df = calculate_ap(_sample_df())
    wb = openpyxl.load_workbook(io.BytesIO(export_excel(df)))
    ws = wb["FMEA Analysis"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "AP" in headers


def test_excel_metadata_includes_ap_counts():
    """W03-3: AP High/Medium/Low tallies appear on the Metadata sheet."""
    df = calculate_ap(_sample_df())
    wb = openpyxl.load_workbook(io.BytesIO(export_excel(df)))
    labels = [wb["Metadata"].cell(r, 1).value for r in range(1, wb["Metadata"].max_row + 1)]
    assert {"AP High", "AP Medium", "AP Low"} <= set(labels)


def test_csv_includes_ap_column_when_present():
    """W03-3: the AP column survives into the CSV export header."""
    df = calculate_ap(_sample_df())
    header = export_csv(df).decode("utf-8").splitlines()[0].split(",")
    assert "AP" in header


def test_pdf_with_ap_column_renders_to_bytes():
    """W03-3: PDF export still produces a valid document with the AP column."""
    df = calculate_ap(_sample_df())
    out = export_pdf(df)
    assert isinstance(out, bytes) and len(out) > 0


def test_exports_omit_ap_when_absent():
    """Back-compat: a pre-AP analyzed frame still exports cleanly (no AP column)."""
    df = _sample_df()  # no AP column
    header = export_csv(df).decode("utf-8").splitlines()[0].split(",")
    assert "AP" not in header
    wb = openpyxl.load_workbook(io.BytesIO(export_excel(df)))
    headers = [wb["FMEA Analysis"].cell(1, c).value for c in range(1, wb["FMEA Analysis"].max_column + 1)]
    assert "AP" not in headers


def test_pdf_export_cleans_tempfile_on_chart_error(monkeypatch, tmp_path):
    """F-009 regression: if a chart-embed step raises during PDF generation,
    we must not leave orphan PNG files in the system temp directory."""
    import os
    import tempfile
    from pathlib import Path

    import pandas as pd

    from fmea_app import exporter
    from fmea_app.rpn_engine import run_pipeline

    # Absolute path so the read works regardless of pytest CWD.
    demo_csv = Path(__file__).parent.parent / "data" / "composite_panel_fmea_demo.csv"
    df = pd.read_csv(demo_csv)
    df = run_pipeline(df)

    def boom(*args, **kwargs):
        raise RuntimeError("simulated chart embed failure")

    tmp_root = tempfile.gettempdir()
    before = {f for f in os.listdir(tmp_root) if f.endswith(".png")}

    monkeypatch.setattr(exporter, "_pdf_chart_page_from_file", boom)
    try:
        exporter.export_pdf(df)
    except RuntimeError:
        pass  # expected — testing cleanup, not error handling

    after = {f for f in os.listdir(tmp_root) if f.endswith(".png")}
    leaked = after - before
    assert not leaked, f"PDF export leaked tempfile(s): {leaked}"
