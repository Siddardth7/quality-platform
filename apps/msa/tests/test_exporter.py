"""Tests for msa_app/exporter.py — Gage R&R report export (W08-3).

MSA-specific export config composed over quality_core.io.export. These tests
load the produced .xlsx back with openpyxl and assert structure/values, decode
the CSVs, assert user-derived strings are formula-injection-escaped, and check
PDFs for valid bytes. The shared export machinery itself is tested in
quality-core.
"""

from __future__ import annotations

import io

import openpyxl
import pandas as pd
import pytest
from msa_app.exporter import (
    VERDICT_SENTENCES,
    GageStudyReport,
    _detail_rows,
    export_csv,
    export_excel,
    export_pdf,
    export_results_csv,
    verdict_sentence,
)

# --- Fixtures ----------------------------------------------------------------

STUDY_ROWS = [
    {"part": "P01", "appraiser": "A", "trial": 1, "measurement": 10.05},
    {"part": "P01", "appraiser": "A", "trial": 2, "measurement": 10.02},
    {"part": "P02", "appraiser": "B", "trial": 1, "measurement": 9.98},
    {"part": "P02", "appraiser": "B", "trial": 2, "measurement": 10.01},
]

RESULTS = {
    "ev": 0.03,
    "av": 0.02,
    "grr": 0.036,
    "pv": 0.5,
    "tv": 0.501,
    "pgrr_study": 7.19,
    "pgrr_tolerance": 12.0,
    "ndc": 6,
    "verdict": "Accept",
    "mean": 10.0,
    "n_parts": 2,
    "n_appraisers": 2,
    "n_trials": 2,
    "is_balanced": True,
}


def _report(**overrides: object) -> GageStudyReport:
    results = {**RESULTS, **overrides.pop("results", {})}  # type: ignore[arg-type]
    kwargs = {"usl": 10.5, "lsl": 9.5, **overrides}
    return GageStudyReport(
        study=pd.DataFrame(STUDY_ROWS),
        results=results,
        **kwargs,
    )


def _no_tolerance_report() -> GageStudyReport:
    return GageStudyReport(
        study=pd.DataFrame(STUDY_ROWS),
        results={**RESULTS, "pgrr_tolerance": None},
        usl=None,
        lsl=None,
    )


def _kv_sheet_to_dict(ws) -> dict[str, object]:
    return {row[0].value: row[1].value for row in ws.iter_rows(min_row=1, max_col=2) if row[0].value}


# --- verdict_sentence ----------------------------------------------------------


@pytest.mark.parametrize("verdict", ["Accept", "Marginal", "Reject"])
def test_verdict_sentence_returns_configured_text(verdict):
    assert verdict_sentence(verdict) == VERDICT_SENTENCES[verdict]


def test_verdict_sentence_fallback_for_unknown_verdict():
    assert verdict_sentence("Bogus") == "Unrecognized verdict — review the study inputs."


# --- _detail_rows --------------------------------------------------------------


def test_detail_rows_tolerance_present():
    rows = dict(_detail_rows(_report()))
    assert rows["%GRR (Tolerance)"] == "12.00%"
    assert rows["USL"] == "10.500000"
    assert rows["LSL"] == "9.500000"
    assert rows["Verdict"] == "Accept"
    assert rows["Verdict Interpretation"] == VERDICT_SENTENCES["Accept"]


def test_detail_rows_tolerance_absent():
    rows = dict(_detail_rows(_no_tolerance_report()))
    assert rows["%GRR (Tolerance)"] == "N/A"
    assert rows["USL"] == "N/A"
    assert rows["LSL"] == "N/A"


# --- export_csv (study data) ----------------------------------------------------


def test_export_csv_header_and_row_count():
    data = export_csv(_report())
    frame = pd.read_csv(io.BytesIO(data))
    assert list(frame.columns) == ["part", "appraiser", "trial", "measurement"]
    assert len(frame) == len(STUDY_ROWS)


def test_export_csv_escapes_formula_injection_in_part():
    report = GageStudyReport(
        study=pd.DataFrame(
            [{"part": "=cmd", "appraiser": "A", "trial": 1, "measurement": 10.0}]
        ),
        results=RESULTS,
        usl=None,
        lsl=None,
    )
    data = export_csv(report)
    text = data.decode("utf-8")
    assert "'=cmd" in text


def test_export_csv_escapes_formula_injection_in_appraiser():
    report = GageStudyReport(
        study=pd.DataFrame(
            [{"part": "P01", "appraiser": "+SUM(A1)", "trial": 1, "measurement": 10.0}]
        ),
        results=RESULTS,
        usl=None,
        lsl=None,
    )
    data = export_csv(report)
    text = data.decode("utf-8")
    assert "'+SUM(A1)" in text


# --- export_results_csv (R1, flat metrics) --------------------------------------


def test_export_results_csv_contains_expected_columns_and_values():
    data = export_results_csv(_report())
    frame = pd.read_csv(io.BytesIO(data))
    assert list(frame.columns) == [
        "EV",
        "AV",
        "GRR",
        "PV",
        "TV",
        "%GRR Study",
        "%GRR Tolerance",
        "ndc",
        "Verdict",
        "Verdict Interpretation",
    ]
    assert len(frame) == 1
    row = frame.iloc[0]
    assert row["ndc"] == 6
    assert row["Verdict"] == "Accept"
    assert row["%GRR Tolerance"] == "12.00%"


def test_export_results_csv_reports_na_when_tolerance_absent():
    data = export_results_csv(_no_tolerance_report())
    frame = pd.read_csv(io.BytesIO(data), na_filter=False)
    assert frame.iloc[0]["%GRR Tolerance"] == "N/A"


@pytest.mark.parametrize("verdict", ["Accept", "Marginal", "Reject"])
def test_export_results_csv_verdict_interpretation_per_verdict(verdict):
    data = export_results_csv(_report(results={"verdict": verdict}))
    frame = pd.read_csv(io.BytesIO(data))
    assert frame.iloc[0]["Verdict Interpretation"] == VERDICT_SENTENCES[verdict]


def test_export_results_csv_fallback_for_unknown_verdict():
    data = export_results_csv(_report(results={"verdict": "Bogus"}))
    frame = pd.read_csv(io.BytesIO(data))
    assert frame.iloc[0]["Verdict Interpretation"] == "Unrecognized verdict — review the study inputs."


# --- export_excel ----------------------------------------------------------------


def test_export_excel_structure_and_values():
    data = export_excel(_report())
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Summary", "Study Data"]

    summary = _kv_sheet_to_dict(wb["Summary"])
    assert summary["Verdict"] == "Accept"
    assert summary["%GRR (Study)"] == "7.19%"
    assert summary["%GRR (Tolerance)"] == "12.00%"
    assert summary["Tool Version"]

    data_ws = wb["Study Data"]
    assert [c.value for c in data_ws[1]] == ["part", "appraiser", "trial", "measurement"]
    assert data_ws.max_row == 1 + len(STUDY_ROWS)


def test_export_excel_reports_na_when_tolerance_and_limits_absent():
    wb = openpyxl.load_workbook(io.BytesIO(export_excel(_no_tolerance_report())))
    summary = _kv_sheet_to_dict(wb["Summary"])
    assert summary["%GRR (Tolerance)"] == "N/A"
    assert summary["USL"] == "N/A"
    assert summary["LSL"] == "N/A"


def test_export_excel_escapes_injection_in_study_data():
    report = GageStudyReport(
        study=pd.DataFrame(
            [{"part": "=cmd", "appraiser": "A", "trial": 1, "measurement": 10.0}]
        ),
        results=RESULTS,
        usl=None,
        lsl=None,
    )
    wb = openpyxl.load_workbook(io.BytesIO(export_excel(report)))
    data_ws = wb["Study Data"]
    assert str(data_ws.cell(row=2, column=1).value).startswith("'=")


# --- export_pdf --------------------------------------------------------------------


def test_export_pdf_is_valid_bytes():
    data = export_pdf(_report())
    assert data.startswith(b"%PDF")
    assert len(data) > 500


def test_export_pdf_handles_tolerance_absent():
    data = export_pdf(_no_tolerance_report())
    assert data.startswith(b"%PDF")
