"""
tests/test_action_export.py
Action-tracking columns in FMEA exports (W05-4b).

When the relational model carries actions, `relational_to_dataframe` appends the
`ACTION_COLUMNS` (blank for rows without an action) and the exporters render them:
Excel/CSV as extra columns, PDF as a dedicated "Action Tracking" page. When no
link has an action, the output is unchanged from the flat path. Owner text stays
formula-injection-safe.
"""

from __future__ import annotations

import io
import re
from datetime import date

import openpyxl
import pandas as pd
from quality_core.io import export_csv
from quality_core.schema import (
    Action,
    ActionStatus,
    FMEADataset,
    FMEARow,
    RelationalFMEA,
    flat_to_relational,
)

from fmea_app.exporter import export_excel, export_pdf
from fmea_app.rpn_engine import ACTION_COLUMNS, run_pipeline_relational

_ROWS = [
    dict(ID=1, Process_Step="Mix", Component="Resin", Function="Bond layers",
         Failure_Mode="Uncured", Effect="Delamination", Severity=9,
         Cause="Low temperature", Occurrence=8, Current_Control="Oven", Detection=5),
    dict(ID=2, Process_Step="Seal", Component="Edge", Function="Seal edge",
         Failure_Mode="Void", Effect="Leak", Severity=6,
         Cause="Gap", Occurrence=3, Current_Control="Visual", Detection=7),
]


def _dataset() -> FMEADataset:
    return FMEADataset(rows=[FMEARow(**r) for r in _ROWS])  # type: ignore[arg-type]


def _model_with_action(owner: str = "Quality Eng") -> RelationalFMEA:
    """A model whose first row (ID 1) carries an action dropping occurrence to 1."""
    model = flat_to_relational(_dataset())
    model.functions[0].failure_modes[0].links[0].action = Action(
        owner=owner, status=ActionStatus.CLOSED, due=date(2026, 8, 1), o_after=1
    )
    return RelationalFMEA(functions=model.functions)  # revalidate with the action


def _pdf_page_count(pdf: bytes) -> int:
    match = re.search(rb"/Count\s+(\d+)", pdf)
    assert match is not None
    return int(match.group(1))


# --- relational_to_dataframe / effectiveness columns --------------------------


def test_action_columns_present_and_correct() -> None:
    df = run_pipeline_relational(_model_with_action())
    assert all(c in df.columns for c in ACTION_COLUMNS)
    row1 = df[df["ID"] == 1].iloc[0]
    assert row1["RPN"] == 360 and row1["RPN_Revised"] == 45 and row1["RPN_Delta"] == -315
    assert row1["AP_Revised"] == "Low"
    assert row1["Action_Status"] == "Closed" and row1["Action_Due"] == "2026-08-01"
    assert row1["O_After"] == 1
    # a row without an action gets blanks, not scores
    row2 = df[df["ID"] == 2].iloc[0]
    assert row2["Action_Owner"] == "" and row2["RPN_Revised"] == ""


def test_no_action_means_no_action_columns() -> None:
    df = run_pipeline_relational(flat_to_relational(_dataset()))
    assert not any(c in df.columns for c in ACTION_COLUMNS)


# --- exporters ----------------------------------------------------------------


def test_excel_includes_action_columns() -> None:
    df = run_pipeline_relational(_model_with_action())
    ws = openpyxl.load_workbook(io.BytesIO(export_excel(df)))["FMEA Analysis"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert "Action_Owner" in headers and "RPN_Revised" in headers and "AP_Revised" in headers


def test_excel_action_owner_is_injection_safe() -> None:
    df = run_pipeline_relational(_model_with_action(owner="=cmd()|calc"))
    ws = openpyxl.load_workbook(io.BytesIO(export_excel(df)))["FMEA Analysis"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    owner_col = headers.index("Action_Owner")
    row1 = next(r for r in ws.iter_rows() if str(r[0].value) == "1")
    cell = row1[owner_col]
    assert cell.data_type != "f"  # never stored as a formula
    assert str(cell.value).startswith("'=")  # escaped


def test_csv_includes_action_columns() -> None:
    df = run_pipeline_relational(_model_with_action())
    header = export_csv(df).decode().splitlines()[0]
    assert "Action_Owner" in header and "RPN_Delta" in header


def test_pdf_adds_action_page_only_when_actions_present() -> None:
    with_actions = export_pdf(run_pipeline_relational(_model_with_action()))
    without = export_pdf(run_pipeline_relational(flat_to_relational(_dataset())))
    assert with_actions[:4] == b"%PDF" and without[:4] == b"%PDF"
    assert _pdf_page_count(with_actions) == _pdf_page_count(without) + 1
